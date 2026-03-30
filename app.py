import os
from pathlib import Path

from flask import Flask, render_template, request, jsonify, send_file

# Heavy deps (video/image/captions) — disabled on cloud via CLOUD_MODE env var
_CLOUD_MODE = os.environ.get("CLOUD_MODE", "").lower() in ("1", "true", "yes")

if not _CLOUD_MODE:
    try:
        from video_processor import start_job, get_job, get_video_info, UPLOADS_DIR, OUTPUT_DIR
        from tiktok_downloader import start_tiktok_job, get_tiktok_job, DOWNLOADS_DIR
        from image_processor import start_image_job, get_image_job
        from captions_processor import (
            start_captions_job, get_captions_job,
            start_captions_batch, get_captions_batch,
            start_reburn_job,
            TEMPLATES as CAPTION_TEMPLATES,
        )
        from voice_generator import start_vo_job, get_vo_job, VOICES as VO_VOICES
        from platform_downloader import start_platform_job, get_platform_job
        from audio_extractor import start_audio_job, get_audio_job
        from video_trimmer import start_trim_job, get_trim_job, get_video_duration
        from bg_remover import start_bg_job, get_bg_job
        _HAS_MEDIA = True
    except ImportError as _e:
        print(f"[WARN] Media import failed: {_e}")
        _HAS_MEDIA = False
        UPLOADS_DIR = Path("uploads")
        OUTPUT_DIR = Path("output")
        DOWNLOADS_DIR = Path("downloads")
else:
    _HAS_MEDIA = False
    UPLOADS_DIR = Path("uploads")
    OUTPUT_DIR = Path("output")
    DOWNLOADS_DIR = Path("downloads")

# Lightweight deps (scraper/tracker) — always available
from shopify_scraper import (
    fetch_collections, start_scrape_job, get_scrape_job,
    start_analyse_job, get_analyse_job,
)
from inventory_tracker import (
    add_tracked_store, get_tracked_stores, remove_tracked_store,
    start_scan_job, get_scan_job, get_store_sales_data,
    start_background_scanner,
)

app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 4 * 1024 * 1024 * 1024  # 4 GB max upload

# Ensure dirs exist
UPLOADS_DIR.mkdir(exist_ok=True)
OUTPUT_DIR.mkdir(exist_ok=True)
DOWNLOADS_DIR.mkdir(exist_ok=True)

# ── Config file for API keys (saved via the Settings page) ──────────────────
import json as _json

CONFIG_FILE = Path(__file__).parent / "config.json"


def _load_config() -> dict:
    try:
        return _json.loads(CONFIG_FILE.read_text())
    except Exception:
        return {}


def _save_config(data: dict):
    CONFIG_FILE.write_text(_json.dumps(data, indent=2))


def _get_api_key(name: str) -> str:
    """Return key from env var first, then from saved config."""
    return os.environ.get(name, "") or _load_config().get(name, "")

# Set output dir for scraper
import shopify_scraper
shopify_scraper.OUTPUT_DIR = OUTPUT_DIR


@app.route("/")
def index():
    return render_template("index.html", has_media=_HAS_MEDIA)


# ── Settings routes ──────────────────────────────────────────────────────────

@app.route("/settings/get")
def settings_get():
    cfg = _load_config()
    openai_key = os.environ.get("OPENAI_API_KEY", "") or cfg.get("OPENAI_API_KEY", "")
    anthropic_key = os.environ.get("ANTHROPIC_API_KEY", "") or cfg.get("ANTHROPIC_API_KEY", "")
    # Mask keys: show only first 8 chars + ****
    def mask(k):
        if not k:
            return ""
        return k[:8] + "****" + k[-4:] if len(k) > 12 else k[:4] + "****"
    return jsonify({
        "openai_key": mask(openai_key),
        "openai_set": bool(openai_key),
        "anthropic_key": mask(anthropic_key),
        "anthropic_set": bool(anthropic_key),
    })


@app.route("/settings/save", methods=["POST"])
def settings_save():
    data = request.get_json() or {}
    cfg = _load_config()
    for key in ("OPENAI_API_KEY", "ANTHROPIC_API_KEY"):
        val = data.get(key, "").strip()
        if val:
            cfg[key] = val
        elif data.get(f"clear_{key}"):
            cfg.pop(key, None)
    _save_config(cfg)
    return jsonify({"ok": True})


@app.before_request
def _check_media_routes():
    if not _HAS_MEDIA:
        path = request.path
        media_prefixes = ("/upload", "/status/", "/download/", "/tiktok", "/image/", "/captions/")
        if any(path.startswith(p) for p in media_prefixes):
            return jsonify({"error": "Media processing unavailable on this server (deploy locally for video/image features)"}), 503


@app.route("/upload", methods=["POST"])
def upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".mp4"):
        return jsonify({"error": "Only MP4 files are accepted"}), 400

    # Save uploaded file
    import uuid
    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}_{file.filename}"
    filepath = UPLOADS_DIR / safe_name
    file.save(str(filepath))

    # Get processing options
    mode = request.form.get("mode", "ffmpeg")  # "ffmpeg" or "ai"
    scale = int(request.form.get("scale", 2))  # 1, 2, or 4
    fps_boost = int(request.form.get("fps_boost", 1))  # 1, 2, or 4

    # Start processing job
    job_id = start_job(filepath, mode=mode, scale=scale, fps_boost=fps_boost)

    return jsonify({"job_id": job_id})


@app.route("/status/<job_id>")
def status(job_id):
    job = get_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    response = {
        "status": job["status"],
        "progress": job["progress"],
        "error": job["error"],
        "mode": job.get("mode", "ffmpeg"),
        "status_message": job.get("status_message", ""),
    }

    if job["input_info"]:
        response["input_info"] = job["input_info"]

    if job["output_info"]:
        response["output_info"] = job["output_info"]
        response["compression_ratio"] = job["compression_ratio"]

    return jsonify(response)


@app.route("/download/<job_id>")
def download(job_id):
    job = get_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404

    output_path = Path(job["output_file"])
    return send_file(
        str(output_path),
        as_attachment=True,
        download_name=f"enhanced_{output_path.name}",
        mimetype="video/mp4"
    )


@app.route("/tiktok", methods=["POST"])
def tiktok():
    data = request.get_json()
    if not data or "urls" not in data:
        return jsonify({"error": "No URLs provided"}), 400

    urls = data["urls"]
    if isinstance(urls, str):
        # Split by newlines, commas, or spaces
        import re
        urls = [u.strip() for u in re.split(r'[\n,]+', urls) if u.strip()]

    enhance_mode = data.get("mode", "ai")
    enhance_scale = int(data.get("scale", 2))
    fps_boost = int(data.get("fps_boost", 1))

    job_id, error = start_tiktok_job(urls, enhance_mode=enhance_mode, enhance_scale=enhance_scale, fps_boost=fps_boost)
    if error:
        return jsonify({"error": error}), 400

    return jsonify({"job_id": job_id})


@app.route("/tiktok/status/<job_id>")
def tiktok_status(job_id):
    job = get_tiktok_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "current_index": job.get("current_index", 0),
        "status_message": job.get("status_message", ""),
        "results": job.get("results", []),
        "error": job.get("error"),
    })


@app.route("/tiktok/download/<job_id>/<int:index>")
def tiktok_download(job_id, index):
    job = get_tiktok_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    results = job.get("results", [])
    if index < 0 or index >= len(results):
        return jsonify({"error": "Invalid index"}), 404

    result = results[index]
    if result["status"] != "done" or not result.get("output_file"):
        return jsonify({"error": "File not ready"}), 404

    output_path = Path(result["output_file"])
    title = result.get("title", "tiktok")
    # Sanitize title for filename
    safe_title = "".join(c for c in title if c.isalnum() or c in " _-")[:50].strip()
    if not safe_title:
        safe_title = "tiktok_video"

    return send_file(
        str(output_path),
        as_attachment=True,
        download_name=f"enhanced_{safe_title}.mp4",
        mimetype="video/mp4"
    )


@app.route("/image/upload", methods=["POST"])
def image_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No filename"}), 400

    ext = file.filename.lower().rsplit(".", 1)[-1] if "." in file.filename else ""
    if ext not in ("png", "jpg", "jpeg", "webp", "bmp", "tiff", "tif"):
        return jsonify({"error": "Format non supporte. Utilisez PNG, JPEG ou WebP."}), 400

    import uuid
    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}_{file.filename}"
    filepath = UPLOADS_DIR / safe_name
    file.save(str(filepath))

    scale = int(request.form.get("scale", 2))
    job_id = start_image_job(filepath, scale=scale)

    return jsonify({"job_id": job_id})


@app.route("/image/status/<job_id>")
def image_status(job_id):
    job = get_image_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    response = {
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "error": job.get("error"),
    }

    if job.get("input_info"):
        response["input_info"] = job["input_info"]
    if job.get("output_info"):
        response["output_info"] = job["output_info"]

    return jsonify(response)


@app.route("/image/download/<job_id>")
def image_download(job_id):
    job = get_image_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404

    output_path = Path(job["output_file"])
    return send_file(
        str(output_path),
        as_attachment=True,
        download_name=f"enhanced_{output_path.name}",
        mimetype="image/png"
    )


@app.route("/captions/upload", methods=["POST"])
def captions_upload():
    files = request.files.getlist("files")
    if not files or all(not f.filename for f in files):
        # Fallback: single file field
        if "file" in request.files:
            files = [request.files["file"]]
        else:
            return jsonify({"error": "No files provided"}), 400

    import uuid as _uuid
    language = request.form.get("language", "auto")
    template = request.form.get("template", "classic")

    saved_paths = []
    for file in files:
        if not file.filename or not file.filename.lower().endswith(".mp4"):
            continue
        file_id = str(_uuid.uuid4())[:8]
        safe_name = f"{file_id}_{file.filename}"
        filepath = UPLOADS_DIR / safe_name
        file.save(str(filepath))
        saved_paths.append(filepath)

    if not saved_paths:
        return jsonify({"error": "Aucun fichier MP4 valide"}), 400

    if len(saved_paths) == 1:
        job_id = start_captions_job(saved_paths[0], language=language, template=template)
        return jsonify({"job_id": job_id, "mode": "single"})
    else:
        batch_id = start_captions_batch(saved_paths, language=language, template=template)
        return jsonify({"job_id": batch_id, "mode": "batch"})


@app.route("/captions/status/<job_id>")
def captions_status(job_id):
    job = get_captions_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "detected_language": job.get("detected_language"),
        "segments_count": job.get("segments_count", 0),
        "output_size_mb": job.get("output_size_mb", 0),
        "error": job.get("error"),
    })


@app.route("/captions/download/<job_id>")
def captions_download(job_id):
    job = get_captions_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404

    output_path = Path(job["output_file"])
    return send_file(
        str(output_path),
        as_attachment=True,
        download_name=f"captioned_{output_path.name}",
        mimetype="video/mp4"
    )


@app.route("/captions/batch/status/<batch_id>")
def captions_batch_status(batch_id):
    batch = get_captions_batch(batch_id)
    if not batch:
        return jsonify({"error": "Batch not found"}), 404

    return jsonify({
        "status": batch["status"],
        "progress": batch["progress"],
        "total": batch["total"],
        "current_index": batch.get("current_index", 0),
        "status_message": batch.get("status_message", ""),
        "results": batch.get("results", []),
        "error": batch.get("error"),
    })


@app.route("/captions/batch/download/<batch_id>/<int:index>")
def captions_batch_download(batch_id, index):
    batch = get_captions_batch(batch_id)
    if not batch:
        return jsonify({"error": "Batch not found"}), 404

    results = batch.get("results", [])
    if index < 0 or index >= len(results):
        return jsonify({"error": "Invalid index"}), 404

    result = results[index]
    if result["status"] != "done" or not result.get("output_file"):
        return jsonify({"error": "File not ready"}), 404

    output_path = Path(result["output_file"])
    filename = result.get("filename", "video")
    safe_name = Path(filename).stem
    return send_file(
        str(output_path),
        as_attachment=True,
        download_name=f"captioned_{safe_name}.mp4",
        mimetype="video/mp4"
    )


@app.route("/captions/output/<job_id>")
def captions_output_video(job_id):
    """Stream the output (burned captions) video for preview."""
    if not _HAS_MEDIA:
        return jsonify({"error": "Media not available"}), 503
    job = get_captions_job(job_id)
    if not job or not job.get("output_file"):
        return jsonify({"error": "Not ready"}), 404
    output_path = Path(job["output_file"])
    if not output_path.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(output_path), mimetype="video/mp4")


@app.route("/voiceover/generate", methods=["POST"])
def voiceover_generate():
    if not _HAS_MEDIA:
        return jsonify({"error": "Media not available"}), 503
    data = request.get_json()
    if not data or not data.get("text", "").strip():
        return jsonify({"error": "Texte vide"}), 400
    text = data["text"].strip()
    if len(text) > 5000:
        return jsonify({"error": "Texte trop long (max 5000 caractères)"}), 400
    voice = data.get("voice", "fr-FR-DeniseNeural")
    rate = max(-50, min(50, int(data.get("rate", 0))))
    pitch = max(-20, min(20, int(data.get("pitch", 0))))
    job_id = start_vo_job(text, voice=voice, rate=rate, pitch=pitch)
    return jsonify({"job_id": job_id})


@app.route("/voiceover/status/<job_id>")
def voiceover_status(job_id):
    if not _HAS_MEDIA:
        return jsonify({"error": "Media not available"}), 503
    job = get_vo_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({k: v for k, v in job.items() if k != "text"})


@app.route("/voiceover/download/<job_id>")
def voiceover_download(job_id):
    if not _HAS_MEDIA:
        return jsonify({"error": "Media not available"}), 503
    job = get_vo_job(job_id)
    if not job or job["status"] != "done" or not job.get("output_file"):
        return jsonify({"error": "Not ready"}), 404
    return send_file(
        str(job["output_file"]),
        as_attachment=True,
        download_name="voiceover.mp3",
        mimetype="audio/mpeg",
    )


@app.route("/voiceover/voices")
def voiceover_voices():
    if not _HAS_MEDIA:
        return jsonify({})
    return jsonify(VO_VOICES)


@app.route("/voiceover/script", methods=["POST"])
def voiceover_script():
    """Generate a voiceover script using OpenAI or Anthropic."""
    import os
    openai_key = _get_api_key("OPENAI_API_KEY")
    anthropic_key = _get_api_key("ANTHROPIC_API_KEY")

    if not openai_key and not anthropic_key:
        return jsonify({"error": "Clé API manquante. Va dans Paramètres et entre ta clé OpenAI ou Anthropic."}), 400

    data = request.get_json()
    if not data or not data.get("prompt", "").strip():
        return jsonify({"error": "Prompt vide"}), 400

    prompt = data["prompt"].strip()
    tone = data.get("tone", "naturel")
    duration = int(data.get("duration", 30))
    language = data.get("language", "français")
    target_words = max(20, int(duration / 60 * 130))

    system = (
        "Tu es un expert en rédaction de scripts pour voix off professionnelles. "
        "Tu génères des textes fluides, naturels et optimisés pour être lus à voix haute. "
        "Pas de titres, pas de listes, pas de formatage Markdown — uniquement du texte narratif continu. "
        "Respecte toujours la langue demandée."
    )
    user_msg = (
        f"Génère un script de voix off d'environ {target_words} mots "
        f"({duration} secondes à rythme naturel).\n"
        f"Sujet / contexte : {prompt}\n"
        f"Ton : {tone}\n"
        f"Langue : {language}\n\n"
        "Réponds uniquement avec le texte du script, rien d'autre."
    )

    try:
        if openai_key:
            from openai import OpenAI
            client = OpenAI(api_key=openai_key)
            response = client.chat.completions.create(
                model="gpt-4o-mini",
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": user_msg},
                ],
                max_tokens=1024,
                temperature=0.8,
            )
            script = response.choices[0].message.content.strip()
        else:
            import anthropic
            client = anthropic.Anthropic(api_key=anthropic_key)
            message = client.messages.create(
                model="claude-3-5-haiku-20241022",
                max_tokens=1024,
                messages=[{"role": "user", "content": user_msg}],
                system=system,
            )
            script = message.content[0].text.strip()

        return jsonify({"script": script, "word_count": len(script.split())})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/captions/input/<job_id>")
def captions_input_video(job_id):
    if not _HAS_MEDIA:
        return jsonify({"error": "Media not available"}), 503
    job = get_captions_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    input_path = Path(job["input_file"])
    if not input_path.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(input_path), mimetype="video/mp4")


@app.route("/captions/segments/<job_id>")
def captions_segments(job_id):
    if not _HAS_MEDIA:
        return jsonify({"error": "Media not available"}), 503
    job = get_captions_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({"segments": job.get("segments", [])})


@app.route("/captions/reburn/<job_id>", methods=["POST"])
def captions_reburn(job_id):
    if not _HAS_MEDIA:
        return jsonify({"error": "Media not available"}), 503
    orig_job = get_captions_job(job_id)
    if not orig_job:
        return jsonify({"error": "Job not found"}), 404
    data = request.get_json()
    edited_segments = data.get("segments", [])
    style = data.get("style", {})
    try:
        new_job_id = start_reburn_job(job_id, edited_segments, style=style)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    return jsonify({"job_id": new_job_id})


@app.route("/platform", methods=["POST"])
def platform_download():
    data = request.get_json()
    if not data or "urls" not in data:
        return jsonify({"error": "No URLs provided"}), 400
    urls = data["urls"]
    if isinstance(urls, str):
        import re
        urls = [u.strip() for u in re.split(r'[\n,]+', urls) if u.strip()]
    job_id, error = start_platform_job(urls)
    if error:
        return jsonify({"error": error}), 400
    return jsonify({"job_id": job_id})


@app.route("/platform/status/<job_id>")
def platform_status(job_id):
    job = get_platform_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "total": job["total"],
        "current_index": job.get("current_index", 0),
        "status_message": job.get("status_message", ""),
        "results": job.get("results", []),
        "error": job.get("error"),
    })


@app.route("/platform/download/<job_id>/<int:index>")
def platform_download_file(job_id, index):
    job = get_platform_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    results = job.get("results", [])
    if index < 0 or index >= len(results):
        return jsonify({"error": "Invalid index"}), 404
    result = results[index]
    if result["status"] != "done" or not result.get("output_file"):
        return jsonify({"error": "File not ready"}), 404
    output_path = Path(result["output_file"])
    title = result.get("title", "video")
    safe_title = "".join(c for c in title if c.isalnum() or c in " _-")[:50].strip() or "video"
    return send_file(str(output_path), as_attachment=True, download_name=f"{safe_title}.mp4", mimetype="video/mp4")


@app.route("/audio/upload", methods=["POST"])
def audio_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".mp4"):
        return jsonify({"error": "Only MP4 files accepted"}), 400
    import uuid
    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}_{file.filename}"
    filepath = UPLOADS_DIR / safe_name
    file.save(str(filepath))
    fmt = request.form.get("format", "mp3")
    job_id = start_audio_job(filepath, fmt=fmt)
    return jsonify({"job_id": job_id})


@app.route("/audio/status/<job_id>")
def audio_status(job_id):
    job = get_audio_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "size_mb": job.get("size_mb"),
        "filename": job.get("filename"),
        "error": job.get("error"),
    })


@app.route("/audio/download/<job_id>")
def audio_download(job_id):
    job = get_audio_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404
    output_path = Path(job["output_file"])
    fmt = job.get("format", "mp3")
    mime = "audio/mpeg" if fmt == "mp3" else "audio/wav"
    return send_file(str(output_path), as_attachment=True, download_name=output_path.name, mimetype=mime)


@app.route("/trim/upload", methods=["POST"])
def trim_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files["file"]
    if not file.filename or not file.filename.lower().endswith(".mp4"):
        return jsonify({"error": "Only MP4 files accepted"}), 400
    import uuid
    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}_{file.filename}"
    filepath = UPLOADS_DIR / safe_name
    file.save(str(filepath))
    duration = get_video_duration(filepath)
    return jsonify({"duration": duration, "filepath": str(filepath)})


@app.route("/trim/start", methods=["POST"])
def trim_start():
    data = request.get_json()
    if not data or "filepath" not in data:
        return jsonify({"error": "No filepath"}), 400
    filepath = data["filepath"]
    start_sec = float(data.get("start", 0))
    end_sec = float(data.get("end", 0))
    if end_sec <= start_sec:
        return jsonify({"error": "end doit être supérieur à start"}), 400
    job_id = start_trim_job(filepath, start_sec, end_sec)
    return jsonify({"job_id": job_id})


@app.route("/trim/status/<job_id>")
def trim_status(job_id):
    job = get_trim_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "size_mb": job.get("size_mb"),
        "error": job.get("error"),
    })


@app.route("/trim/download/<job_id>")
def trim_download(job_id):
    job = get_trim_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404
    output_path = Path(job["output_file"])
    return send_file(str(output_path), as_attachment=True, download_name=output_path.name, mimetype="video/mp4")


@app.route("/bgremove/upload", methods=["POST"])
def bgremove_upload():
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400
    file = request.files["file"]
    if not file.filename:
        return jsonify({"error": "No filename"}), 400
    ext = file.filename.lower().rsplit(".", 1)[-1] if "." in file.filename else ""
    if ext not in ("png", "jpg", "jpeg", "webp", "bmp"):
        return jsonify({"error": "Format non supporté"}), 400
    import uuid
    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}_{file.filename}"
    filepath = UPLOADS_DIR / safe_name
    file.save(str(filepath))
    job_id = start_bg_job(filepath)
    return jsonify({"job_id": job_id})


@app.route("/bgremove/status/<job_id>")
def bgremove_status(job_id):
    job = get_bg_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "size_kb": job.get("size_kb"),
        "error": job.get("error"),
    })


@app.route("/bgremove/download/<job_id>")
def bgremove_download(job_id):
    job = get_bg_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404
    output_path = Path(job["output_file"])
    return send_file(str(output_path), as_attachment=True, download_name=output_path.name, mimetype="image/png")


@app.route("/scraper/collections", methods=["POST"])
def scraper_collections():
    data = request.get_json()
    if not data or not data.get("url"):
        return jsonify({"error": "URL requise"}), 400

    try:
        collections = fetch_collections(data["url"])
        return jsonify({"collections": collections})
    except Exception as e:
        return jsonify({"error": str(e)}), 400


@app.route("/scraper/start", methods=["POST"])
def scraper_start():
    data = request.get_json()
    if not data or not data.get("url"):
        return jsonify({"error": "URL requise"}), 400

    store_url = data["url"]
    collections = data.get("collections", [])  # list of handles

    job_id = start_scrape_job(store_url, collections=collections)
    return jsonify({"job_id": job_id})


@app.route("/scraper/status/<job_id>")
def scraper_status(job_id):
    job = get_scrape_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "products_count": job.get("products_count", 0),
        "rows_count": job.get("rows_count", 0),
        "error": job.get("error"),
    })


@app.route("/scraper/download/<job_id>")
def scraper_download(job_id):
    job = get_scrape_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404

    output_path = Path(job["output_file"])
    # Use store domain for filename
    from urllib.parse import urlparse
    domain = urlparse(job["store_url"]).netloc.replace("www.", "")
    safe_domain = "".join(c for c in domain if c.isalnum() or c in ".-_")

    return send_file(
        str(output_path),
        as_attachment=True,
        download_name=f"products_{safe_domain}.csv",
        mimetype="text/csv"
    )


@app.route("/scraper/download/<job_id>/excel")
def scraper_download_excel(job_id):
    job = get_scrape_job(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "File not ready"}), 404

    csv_path = Path(job["output_file"])
    from urllib.parse import urlparse
    domain = urlparse(job["store_url"]).netloc.replace("www.", "")
    safe_domain = "".join(c for c in domain if c.isalnum() or c in ".-_")
    excel_path = csv_path.with_suffix(".xlsx")

    try:
        import csv
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Produits"

        header_fill = PatternFill("solid", fgColor="6D28D9")
        header_font = Font(bold=True, color="FFFFFF")

        with open(str(csv_path), encoding="utf-8") as f:
            reader = csv.reader(f)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if row_idx == 1:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")

        # Auto width
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 50)

        wb.save(str(excel_path))
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return send_file(
        str(excel_path),
        as_attachment=True,
        download_name=f"products_{safe_domain}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/analyse/start", methods=["POST"])
def analyse_start():
    data = request.get_json()
    if not data or not data.get("url"):
        return jsonify({"error": "URL requise"}), 400

    job_id = start_analyse_job(data["url"])
    return jsonify({"job_id": job_id})


@app.route("/analyse/status/<job_id>")
def analyse_status(job_id):
    job = get_analyse_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404

    response = {
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "error": job.get("error"),
    }

    if job["status"] == "done":
        response["store_name"] = job["store_name"]
        response["total_products"] = job["total_products"]
        response["total_revenue_est"] = job["total_revenue_est"]
        response["products"] = job["products"]

    return jsonify(response)


@app.route("/tracker/stores", methods=["GET"])
def tracker_list_stores():
    stores = get_tracked_stores()
    return jsonify({"stores": stores})


@app.route("/tracker/add", methods=["POST"])
def tracker_add_store():
    data = request.get_json()
    if not data or not data.get("url"):
        return jsonify({"error": "URL requise"}), 400

    store_id = add_tracked_store(data["url"])
    return jsonify({"store_id": store_id})


@app.route("/tracker/remove/<store_id>", methods=["POST"])
def tracker_remove_store(store_id):
    remove_tracked_store(store_id)
    return jsonify({"ok": True})


@app.route("/tracker/scan/<store_id>", methods=["POST"])
def tracker_scan(store_id):
    job_id = start_scan_job(store_id)
    return jsonify({"job_id": job_id})


@app.route("/tracker/scan/status/<job_id>")
def tracker_scan_status(job_id):
    job = get_scan_job(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "status_message": job.get("status_message", ""),
        "error": job.get("error"),
    })


@app.route("/tracker/data/<store_id>")
def tracker_data(store_id):
    days = request.args.get("days", 30, type=int)
    data = get_store_sales_data(store_id, days=days)
    if not data:
        return jsonify({"error": "Store not found"}), 404
    return jsonify(data)


# ── Social media API keys in settings ─────────────────────────────────────────

@app.route("/settings/social/save", methods=["POST"])
def settings_social_save():
    """Save social platform API keys."""
    data = request.get_json() or {}
    cfg = _load_config()
    social_keys = [
        "TIKTOK_CLIENT_KEY", "TIKTOK_CLIENT_SECRET",
        "YOUTUBE_CLIENT_ID", "YOUTUBE_CLIENT_SECRET",
        "INSTAGRAM_APP_ID", "INSTAGRAM_APP_SECRET",
    ]
    for key in social_keys:
        val = data.get(key, "").strip()
        if val and not val.endswith("****"):
            cfg[key] = val
    _save_config(cfg)
    return jsonify({"ok": True})


@app.route("/settings/social/get")
def settings_social_get():
    """Return masked social platform keys status."""
    cfg = _load_config()
    def has(k):
        return bool(os.environ.get(k, "") or cfg.get(k, ""))
    def mask(k):
        v = os.environ.get(k, "") or cfg.get(k, "")
        if not v:
            return ""
        return v[:4] + "****" + v[-4:] if len(v) > 8 else "****"
    return jsonify({
        "tiktok_key_set": has("TIKTOK_CLIENT_KEY"),
        "tiktok_key": mask("TIKTOK_CLIENT_KEY"),
        "youtube_id_set": has("YOUTUBE_CLIENT_ID"),
        "youtube_id": mask("YOUTUBE_CLIENT_ID"),
        "instagram_id_set": has("INSTAGRAM_APP_ID"),
        "instagram_id": mask("INSTAGRAM_APP_ID"),
    })


# ── Automation routes ──────────────────────────────────────────────────────────

import uuid as _uuid_mod
import subprocess
import threading

# In-memory job store for automation clips
_automation_jobs = {}


def _run_clips_job(job_id, source_path, clip_duration, num_clips, style, accent_color):
    """Background worker: cut clips from a video using ffmpeg."""
    job = _automation_jobs[job_id]
    try:
        import subprocess
        from pathlib import Path

        output_dir = OUTPUT_DIR / f"clips_{job_id}"
        output_dir.mkdir(exist_ok=True)

        # Get video duration via ffprobe
        result = subprocess.run(
            ["ffprobe", "-v", "error", "-show_entries", "format=duration",
             "-of", "default=noprint_wrappers=1:nokey=1", str(source_path)],
            capture_output=True, text=True, timeout=30
        )
        total_duration = float(result.stdout.strip())
        job["total_duration"] = total_duration

        clips = []
        step = max(1.0, (total_duration - clip_duration) / max(1, num_clips - 1)) if num_clips > 1 else 0
        start_times = [i * step for i in range(num_clips)]
        # Ensure no clip goes past the end
        start_times = [s for s in start_times if s + clip_duration <= total_duration + 1][:num_clips]

        for i, start in enumerate(start_times):
            clip_path = output_dir / f"clip_{i+1:02d}.mp4"
            cmd = [
                "ffmpeg", "-y",
                "-ss", str(start),
                "-i", str(source_path),
                "-t", str(clip_duration),
                "-c:v", "libx264", "-crf", "23", "-preset", "fast",
                "-c:a", "aac", "-b:a", "128k",
                "-movflags", "+faststart",
                str(clip_path)
            ]
            subprocess.run(cmd, capture_output=True, timeout=300)

            clips.append({
                "index": i,
                "filename": clip_path.name,
                "path": str(clip_path),
                "start": round(start, 1),
                "duration": clip_duration,
                "status": "done" if clip_path.exists() else "error",
                "size_mb": round(clip_path.stat().st_size / 1024 / 1024, 1) if clip_path.exists() else 0,
            })
            job["progress"] = int((i + 1) / len(start_times) * 100)
            job["clips"] = clips[:]

        job["status"] = "done"
        job["clips"] = clips
        job["progress"] = 100
    except Exception as e:
        job["status"] = "error"
        job["error"] = str(e)


@app.route("/automation/clips", methods=["POST"])
def automation_clips():
    """Start a clip generation job."""
    clip_duration = int(request.form.get("clip_duration", 30))
    num_clips = max(1, min(10, int(request.form.get("num_clips", 3))))
    style = request.form.get("style", "bold_pop")
    accent_color = request.form.get("accent_color", "#ffee00")
    url = request.form.get("url", "").strip()

    source_path = None

    # File upload
    if "file" in request.files and request.files["file"].filename:
        file = request.files["file"]
        if not file.filename.lower().endswith(".mp4"):
            return jsonify({"error": "Only MP4 files accepted"}), 400
        file_id = str(_uuid_mod.uuid4())[:8]
        safe_name = f"{file_id}_{file.filename}"
        source_path = UPLOADS_DIR / safe_name
        file.save(str(source_path))
    elif url:
        # Download from URL using yt-dlp if available
        file_id = str(_uuid_mod.uuid4())[:8]
        source_path = UPLOADS_DIR / f"{file_id}_automation_source.mp4"
        try:
            result = subprocess.run(
                ["yt-dlp", "-f", "mp4/best[ext=mp4]", "-o", str(source_path), url],
                capture_output=True, text=True, timeout=120
            )
            if not source_path.exists():
                return jsonify({"error": f"Impossible de télécharger la vidéo: {result.stderr[:200]}"}), 400
        except FileNotFoundError:
            return jsonify({"error": "yt-dlp non installé. Installez-le via: pip install yt-dlp"}), 400
        except Exception as e:
            return jsonify({"error": str(e)}), 400
    else:
        return jsonify({"error": "Fournissez un fichier MP4 ou une URL"}), 400

    job_id = str(_uuid_mod.uuid4())[:8]
    _automation_jobs[job_id] = {
        "status": "processing",
        "progress": 0,
        "clips": [],
        "error": None,
        "clip_duration": clip_duration,
        "num_clips": num_clips,
    }

    t = threading.Thread(
        target=_run_clips_job,
        args=(job_id, source_path, clip_duration, num_clips, style, accent_color),
        daemon=True
    )
    t.start()

    return jsonify({"job_id": job_id})


@app.route("/automation/clips/status/<job_id>")
def automation_clips_status(job_id):
    """Return clip generation job status."""
    job = _automation_jobs.get(job_id)
    if not job:
        return jsonify({"error": "Job not found"}), 404
    return jsonify({
        "status": job["status"],
        "progress": job["progress"],
        "clips": job.get("clips", []),
        "error": job.get("error"),
        "num_clips": job.get("num_clips", 0),
    })


@app.route("/automation/clips/download/<job_id>/<int:index>")
def automation_clips_download(job_id, index):
    """Download a generated clip."""
    job = _automation_jobs.get(job_id)
    if not job or job["status"] != "done":
        return jsonify({"error": "Not ready"}), 404
    clips = job.get("clips", [])
    if index < 0 or index >= len(clips):
        return jsonify({"error": "Invalid index"}), 404
    clip = clips[index]
    clip_path = Path(clip["path"])
    if not clip_path.exists():
        return jsonify({"error": "File not found"}), 404
    return send_file(str(clip_path), as_attachment=True,
                     download_name=clip_path.name, mimetype="video/mp4")


# ── Publish routes ─────────────────────────────────────────────────────────────

@app.route("/publish/status")
def publish_status():
    """Return connected platforms status."""
    cfg = _load_config()
    return jsonify({
        "tiktok": bool(cfg.get("TIKTOK_ACCESS_TOKEN")),
        "youtube": bool(cfg.get("YOUTUBE_ACCESS_TOKEN")),
        "instagram": bool(cfg.get("INSTAGRAM_ACCESS_TOKEN")),
    })


# ── TikTok OAuth ───────────────────────────────────────────────────────────────

@app.route("/publish/tiktok/auth")
def tiktok_auth():
    """Redirect user to TikTok OAuth page."""
    from flask import redirect as flask_redirect
    client_key = _get_api_key("TIKTOK_CLIENT_KEY")
    if not client_key:
        return jsonify({"error": "TIKTOK_CLIENT_KEY non configuré. Va dans Paramètres."}), 400
    redirect_uri = "http://localhost:5050/publish/tiktok/callback"
    scope = "video.publish,video.upload"
    csrf_state = str(_uuid_mod.uuid4())[:16]
    url = (
        f"https://www.tiktok.com/v2/auth/authorize/"
        f"?client_key={client_key}"
        f"&scope={scope}"
        f"&response_type=code"
        f"&redirect_uri={redirect_uri}"
        f"&state={csrf_state}"
    )
    return flask_redirect(url)


@app.route("/publish/tiktok/callback")
def tiktok_callback():
    """Handle TikTok OAuth callback, exchange code for token."""
    code = request.args.get("code")
    error = request.args.get("error")
    if error or not code:
        return f"<h2>Erreur TikTok OAuth</h2><p>{error or 'No code received'}</p>", 400

    client_key = _get_api_key("TIKTOK_CLIENT_KEY")
    client_secret = _get_api_key("TIKTOK_CLIENT_SECRET")
    redirect_uri = "http://localhost:5050/publish/tiktok/callback"

    import requests as _requests
    try:
        resp = _requests.post(
            "https://open.tiktokapis.com/v2/oauth/token/",
            data={
                "client_key": client_key,
                "client_secret": client_secret,
                "code": code,
                "grant_type": "authorization_code",
                "redirect_uri": redirect_uri,
            },
            headers={"Content-Type": "application/x-www-form-urlencoded"},
            timeout=15
        )
        token_data = resp.json()
        if token_data.get("access_token"):
            cfg = _load_config()
            cfg["TIKTOK_ACCESS_TOKEN"] = token_data["access_token"]
            cfg["TIKTOK_REFRESH_TOKEN"] = token_data.get("refresh_token", "")
            cfg["TIKTOK_OPEN_ID"] = token_data.get("open_id", "")
            _save_config(cfg)
            return """<html><body style="background:#080810;color:#e0e0ee;font-family:sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;text-align:center;">
                <div><h2 style="color:#10b981;">✓ TikTok connecté !</h2><p>Vous pouvez fermer cette fenêtre.</p>
                <script>setTimeout(()=>{window.close();},2000);</script></div></body></html>"""
        else:
            return f"<h2>Erreur</h2><p>{token_data}</p>", 400
    except Exception as e:
        return f"<h2>Erreur</h2><p>{str(e)}</p>", 500


@app.route("/publish/tiktok/upload", methods=["POST"])
def tiktok_upload():
    """Upload a video to TikTok using the Content Posting API."""
    cfg = _load_config()
    access_token = cfg.get("TIKTOK_ACCESS_TOKEN")
    if not access_token:
        return jsonify({"error": "TikTok non connecté. Connectez-vous d'abord."}), 401

    data = request.get_json() or {}
    video_path = data.get("video_path", "")
    title = data.get("title", "")
    description = data.get("description", "")
    hashtags = data.get("hashtags", "")

    if not video_path or not Path(video_path).exists():
        return jsonify({"error": "Fichier vidéo introuvable"}), 400

    import requests as _requests
    try:
        file_size = Path(video_path).stat().st_size
        # Step 1: Initialize upload
        init_resp = _requests.post(
            "https://open.tiktokapis.com/v2/post/publish/video/init/",
            headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json; charset=UTF-8"},
            json={
                "post_info": {
                    "title": (title + " " + hashtags)[:150].strip(),
                    "privacy_level": "SELF_ONLY",  # safer default
                    "disable_duet": False,
                    "disable_comment": False,
                    "disable_stitch": False,
                    "video_cover_timestamp_ms": 1000,
                },
                "source_info": {
                    "source": "FILE_UPLOAD",
                    "video_size": file_size,
                    "chunk_size": file_size,
                    "total_chunk_count": 1,
                }
            },
            timeout=30
        )
        init_data = init_resp.json()
        if init_data.get("error", {}).get("code", "ok") != "ok":
            return jsonify({"error": init_data.get("error", {}).get("message", "Init failed")}), 400

        publish_id = init_data.get("data", {}).get("publish_id")
        upload_url = init_data.get("data", {}).get("upload_url")

        if not upload_url:
            return jsonify({"error": "No upload URL received from TikTok"}), 400

        # Step 2: Upload video
        with open(video_path, "rb") as f:
            video_data = f.read()

        upload_resp = _requests.put(
            upload_url,
            data=video_data,
            headers={
                "Content-Type": "video/mp4",
                "Content-Range": f"bytes 0-{file_size - 1}/{file_size}",
                "Content-Length": str(file_size),
            },
            timeout=300
        )

        if upload_resp.status_code not in (200, 201, 206):
            return jsonify({"error": f"Upload failed: HTTP {upload_resp.status_code}"}), 400

        return jsonify({"ok": True, "publish_id": publish_id, "platform": "tiktok"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── YouTube OAuth ──────────────────────────────────────────────────────────────

@app.route("/publish/youtube/auth")
def youtube_auth():
    """Redirect to Google OAuth for YouTube."""
    from flask import redirect as flask_redirect
    client_id = _get_api_key("YOUTUBE_CLIENT_ID")
    if not client_id:
        return jsonify({"error": "YOUTUBE_CLIENT_ID non configuré. Va dans Paramètres."}), 400
    redirect_uri = "http://localhost:5050/publish/youtube/callback"
    scope = "https://www.googleapis.com/auth/youtube.upload"
    url = (
        "https://accounts.google.com/o/oauth2/v2/auth"
        f"?client_id={client_id}"
        f"&redirect_uri={redirect_uri}"
        f"&response_type=code"
        f"&scope={scope}"
        f"&access_type=offline"
        f"&prompt=consent"
    )
    return flask_redirect(url)


@app.route("/publish/youtube/callback")
def youtube_callback():
    """Handle Google OAuth callback for YouTube."""
    code = request.args.get("code")
    error = request.args.get("error")
    if error or not code:
        return f"<h2>Erreur Google OAuth</h2><p>{error or 'No code'}</p>", 400

    client_id = _get_api_key("YOUTUBE_CLIENT_ID")
    client_secret = _get_api_key("YOUTUBE_CLIENT_SECRET")
    redirect_uri = "http://localhost:5050/publish/youtube/callback"

    import requests as _requests
    try:
        resp = _requests.post(
            "https://oauth2.googleapis.com/token",
            data={
                "code": code,
                "client_id": client_id,
                "client_secret": client_secret,
                "redirect_uri": redirect_uri,
                "grant_type": "authorization_code",
            },
            timeout=15
        )
        token_data = resp.json()
        if token_data.get("access_token"):
            cfg = _load_config()
            cfg["YOUTUBE_ACCESS_TOKEN"] = token_data["access_token"]
            cfg["YOUTUBE_REFRESH_TOKEN"] = token_data.get("refresh_token", "")
            _save_config(cfg)
            return """<html><body style="background:#080810;color:#e0e0ee;font-family:sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;text-align:center;">
                <div><h2 style="color:#10b981;">✓ YouTube connecté !</h2><p>Vous pouvez fermer cette fenêtre.</p>
                <script>setTimeout(()=>{window.close();},2000);</script></div></body></html>"""
        else:
            return f"<h2>Erreur</h2><p>{token_data}</p>", 400
    except Exception as e:
        return f"<h2>Erreur</h2><p>{str(e)}</p>", 500


@app.route("/publish/youtube/upload", methods=["POST"])
def youtube_upload():
    """Upload a video to YouTube."""
    cfg = _load_config()
    access_token = cfg.get("YOUTUBE_ACCESS_TOKEN")
    if not access_token:
        return jsonify({"error": "YouTube non connecté. Connectez-vous d'abord."}), 401

    data = request.get_json() or {}
    video_path = data.get("video_path", "")
    title = data.get("title", "Ma vidéo")
    description = data.get("description", "")
    hashtags = data.get("hashtags", "")

    if not video_path or not Path(video_path).exists():
        return jsonify({"error": "Fichier vidéo introuvable"}), 400

    import requests as _requests
    try:
        # Build metadata
        full_desc = description + ("\n\n" + hashtags if hashtags else "")
        metadata = {
            "snippet": {
                "title": title[:100],
                "description": full_desc[:5000],
                "tags": [t.strip("#") for t in hashtags.split() if t.startswith("#")][:15],
                "categoryId": "22"
            },
            "status": {"privacyStatus": "private"}
        }

        # Initiate resumable upload
        init_resp = _requests.post(
            "https://www.googleapis.com/upload/youtube/v3/videos?uploadType=resumable&part=snippet,status",
            headers={
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/json; charset=UTF-8",
                "X-Upload-Content-Type": "video/mp4",
                "X-Upload-Content-Length": str(Path(video_path).stat().st_size),
            },
            json=metadata,
            timeout=30
        )

        if init_resp.status_code not in (200, 201):
            return jsonify({"error": f"YouTube init failed: {init_resp.text[:300]}"}), 400

        upload_url = init_resp.headers.get("Location")
        if not upload_url:
            return jsonify({"error": "No upload URL from YouTube"}), 400

        # Upload file
        with open(video_path, "rb") as f:
            upload_resp = _requests.put(
                upload_url,
                data=f,
                headers={"Content-Type": "video/mp4"},
                timeout=600
            )

        if upload_resp.status_code in (200, 201):
            video_id = upload_resp.json().get("id", "")
            return jsonify({
                "ok": True,
                "platform": "youtube",
                "video_id": video_id,
                "url": f"https://www.youtube.com/watch?v={video_id}"
            })
        else:
            return jsonify({"error": f"YouTube upload failed: {upload_resp.text[:300]}"}), 400
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ── Instagram OAuth ────────────────────────────────────────────────────────────

@app.route("/publish/instagram/auth")
def instagram_auth():
    """Redirect to Meta OAuth for Instagram."""
    from flask import redirect as flask_redirect
    app_id = _get_api_key("INSTAGRAM_APP_ID")
    if not app_id:
        return jsonify({"error": "INSTAGRAM_APP_ID non configuré. Va dans Paramètres."}), 400
    redirect_uri = "http://localhost:5050/publish/instagram/callback"
    scope = "instagram_basic,instagram_content_publish,pages_show_list"
    url = (
        "https://www.facebook.com/v18.0/dialog/oauth"
        f"?client_id={app_id}"
        f"&redirect_uri={redirect_uri}"
        f"&scope={scope}"
        f"&response_type=code"
    )
    return flask_redirect(url)


@app.route("/publish/instagram/callback")
def instagram_callback():
    """Handle Meta OAuth callback for Instagram."""
    code = request.args.get("code")
    error = request.args.get("error")
    if error or not code:
        return f"<h2>Erreur Meta OAuth</h2><p>{error or 'No code'}</p>", 400

    app_id = _get_api_key("INSTAGRAM_APP_ID")
    app_secret = _get_api_key("INSTAGRAM_APP_SECRET")
    redirect_uri = "http://localhost:5050/publish/instagram/callback"

    import requests as _requests
    try:
        resp = _requests.get(
            "https://graph.facebook.com/v18.0/oauth/access_token",
            params={
                "client_id": app_id,
                "redirect_uri": redirect_uri,
                "client_secret": app_secret,
                "code": code,
            },
            timeout=15
        )
        token_data = resp.json()
        if token_data.get("access_token"):
            cfg = _load_config()
            cfg["INSTAGRAM_ACCESS_TOKEN"] = token_data["access_token"]
            _save_config(cfg)
            return """<html><body style="background:#080810;color:#e0e0ee;font-family:sans-serif;display:flex;align-items:center;justify-content:center;height:100vh;text-align:center;">
                <div><h2 style="color:#10b981;">✓ Instagram connecté !</h2><p>Vous pouvez fermer cette fenêtre.</p>
                <script>setTimeout(()=>{window.close();},2000);</script></div></body></html>"""
        else:
            return f"<h2>Erreur</h2><p>{token_data}</p>", 400
    except Exception as e:
        return f"<h2>Erreur</h2><p>{str(e)}</p>", 500


@app.route("/publish/post", methods=["POST"])
def publish_post():
    """
    Publish a video to one or more platforms.
    Accepts JSON: { video_path, title, description, hashtags, platforms: ["tiktok","youtube","instagram"], scheduled_time? }
    """
    data = request.get_json() or {}
    video_path = data.get("video_path", "").strip()
    title = data.get("title", "").strip() or "Ma vidéo"
    description = data.get("description", "").strip()
    hashtags = data.get("hashtags", "").strip()
    platforms = data.get("platforms", [])

    if not video_path:
        return jsonify({"error": "Aucun fichier sélectionné"}), 400
    if not Path(video_path).exists():
        return jsonify({"error": f"Fichier introuvable: {video_path}"}), 400
    if not platforms:
        return jsonify({"error": "Sélectionnez au moins un réseau"}), 400

    cfg = _load_config()
    results = {}

    import requests as _requests

    for platform in platforms:
        if platform == "tiktok":
            if not cfg.get("TIKTOK_ACCESS_TOKEN"):
                results["tiktok"] = {"ok": False, "error": "Non connecté"}
                continue
            # Call our own upload endpoint internally
            with app.test_request_context():
                pass
            try:
                access_token = cfg["TIKTOK_ACCESS_TOKEN"]
                file_size = Path(video_path).stat().st_size
                init_resp = _requests.post(
                    "https://open.tiktokapis.com/v2/post/publish/video/init/",
                    headers={"Authorization": f"Bearer {access_token}", "Content-Type": "application/json; charset=UTF-8"},
                    json={
                        "post_info": {
                            "title": (title + " " + hashtags)[:150].strip(),
                            "privacy_level": "SELF_ONLY",
                            "disable_duet": False, "disable_comment": False, "disable_stitch": False,
                            "video_cover_timestamp_ms": 1000,
                        },
                        "source_info": {"source": "FILE_UPLOAD", "video_size": file_size, "chunk_size": file_size, "total_chunk_count": 1}
                    }, timeout=30
                )
                init_data = init_resp.json()
                upload_url = init_data.get("data", {}).get("upload_url")
                publish_id = init_data.get("data", {}).get("publish_id")
                if upload_url:
                    with open(video_path, "rb") as f:
                        _requests.put(upload_url, data=f.read(), headers={
                            "Content-Type": "video/mp4",
                            "Content-Range": f"bytes 0-{file_size-1}/{file_size}",
                            "Content-Length": str(file_size),
                        }, timeout=300)
                    results["tiktok"] = {"ok": True, "publish_id": publish_id}
                else:
                    results["tiktok"] = {"ok": False, "error": str(init_data)}
            except Exception as e:
                results["tiktok"] = {"ok": False, "error": str(e)}

        elif platform == "youtube":
            if not cfg.get("YOUTUBE_ACCESS_TOKEN"):
                results["youtube"] = {"ok": False, "error": "Non connecté"}
                continue
            try:
                access_token = cfg["YOUTUBE_ACCESS_TOKEN"]
                full_desc = description + ("\n\n" + hashtags if hashtags else "")
                metadata = {
                    "snippet": {"title": title[:100], "description": full_desc[:5000], "categoryId": "22"},
                    "status": {"privacyStatus": "private"}
                }
                init_resp = _requests.post(
                    "https://www.googleapis.com/upload/youtube/v3/videos?uploadType=resumable&part=snippet,status",
                    headers={
                        "Authorization": f"Bearer {access_token}",
                        "Content-Type": "application/json; charset=UTF-8",
                        "X-Upload-Content-Type": "video/mp4",
                        "X-Upload-Content-Length": str(Path(video_path).stat().st_size),
                    },
                    json=metadata, timeout=30
                )
                if init_resp.status_code in (200, 201):
                    upload_url = init_resp.headers.get("Location")
                    with open(video_path, "rb") as f:
                        up = _requests.put(upload_url, data=f, headers={"Content-Type": "video/mp4"}, timeout=600)
                    if up.status_code in (200, 201):
                        video_id = up.json().get("id", "")
                        results["youtube"] = {"ok": True, "url": f"https://www.youtube.com/watch?v={video_id}"}
                    else:
                        results["youtube"] = {"ok": False, "error": up.text[:200]}
                else:
                    results["youtube"] = {"ok": False, "error": init_resp.text[:200]}
            except Exception as e:
                results["youtube"] = {"ok": False, "error": str(e)}

        elif platform == "instagram":
            if not cfg.get("INSTAGRAM_ACCESS_TOKEN"):
                results["instagram"] = {"ok": False, "error": "Non connecté"}
            else:
                results["instagram"] = {"ok": False, "error": "Instagram video posting requires a public video URL. Upload to your server first."}

    # Save to publish history in config
    history = cfg.get("PUBLISH_HISTORY", [])
    import time
    history.insert(0, {
        "title": title,
        "platforms": platforms,
        "results": results,
        "date": _json.dumps({"ts": int(time.time())}),
        "video": Path(video_path).name,
    })
    cfg["PUBLISH_HISTORY"] = history[:50]
    _save_config(cfg)

    return jsonify({"ok": True, "results": results})


@app.route("/publish/history")
def publish_history():
    """Return publication history."""
    cfg = _load_config()
    import time
    history = cfg.get("PUBLISH_HISTORY", [])
    return jsonify({"history": history})


# Start background scanner when running via gunicorn too
start_background_scanner()

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5050))
    app.run(debug=False, host="0.0.0.0", port=port)
